# -*- coding: utf-8 -*-
"""
Created on Wed Sep 18 13:47:29 2019
EXCEL File read and rank with specified order
@author: duqs
Company: DFTC
"""

import os,sys,time

import pandas as pd
import numpy as np

import wx       #   wxPython

#from  openpyxl import Workbook
from openpyxl import load_workbook

# **************************************************************** #

path0 = os.getcwd()
os.chdir(path0)
path1 = path0 + '\\excel_file'


wb = load_workbook('VarList.xlsx')
ws = wb.active

rows = []
for row in ws.iter_rows():
    rows.append(row)

col_max = ws.max_column

var_target = []
for i in range(col_max):
    var_target.append(rows[0][i].value)


app = wx.App()

wildcard = 'excel workbook(*.xlsx)|*.xlsx|excel 97-2003 workbook(*.xls)|*.xls|All Files(*.*)|*.*'
dlg0 = wx.FileDialog(None,message = 'Choose a EXCEL file',
                                     defaultDir = path1,
                                     defaultFile="",
                                     wildcard = wildcard,
                                     style = wx.FD_OPEN|wx.FD_MULTIPLE)
if dlg0.ShowModal() == wx.ID_OK:
        file_path = dlg0.GetPaths() #include path and filename
else:
    time.sleep(1)
    wx.Exit()   #exit wxPython
    sys.exit()

file_num = len(file_path)

def newlist(num):
        newlist = []
        for i in range(num):
            newlist.append([])
        return newlist

df = []
for i in range(file_num):
    df.append(pd.read_excel(file_path[i]))

index = []
for i in range(file_num):
    index.append(df[i].iloc[:1].size)

NAN_list = []
for i in range(file_num):
    for j in range(index[i]):
        NAN_list.append(np.NAN)

for i in range(file_num):
    for j in range(len(var_target)):
        if var_target[j] not in df[i].columns:
            df[i][var_target[j]] = NAN_list[i]

df_filter = []
for i in range(file_num):
    df_filter.append(df[i][var_target])

path2 = 'excel_output\\'
isExists = os.path.exists(path2)
if not isExists:
    os.mkdir(path2)
else:
    pass

excel_name = []
for i in range(file_num):
    num1 = file_path[i].rfind('\\')
    num2 = file_path[i].rfind('.')
    if file_path[i][num2 + 1:] == 'xlsx':
        excel_name.append(file_path[i][num1 + 1:-5]) #delete '.xlsx'
    else:
        excel_name.append(file_path[i][num1 + 1:-4]) #delete '.xls'


#sort according to the sequence in VarList.xlsx
for i in range(file_num):
    df_filter[i].to_excel("excel_output\\%s.xlsx" % excel_name[i],encoding="utf_8",index=False,columns = var_target)

wx.MessageDialog(None,message = 'excel rank Complete',caption = 'Message')

app.MainLoop()
del app









