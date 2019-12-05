# -*- coding: utf-8 -*-
"""
Created on Wed Oct  9 11:19:22 2019

@author: duqs
"""

import os

from openpyxl import Workbook
from openpyxl import load_workbook

import pandas as pd
import numpy as np

import mdfreader

import win32api

import re

from scipy.interpolate import interp1d

import wx
import wx.xrc

# *************************************************************************************** #

class MyWindow ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Data Import and Process", pos = wx.DefaultPosition, size = wx.Size( 700,700 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

        # set the frame icon
        self.icon = wx.Icon('dfm.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(self.icon)

        # the outmost boxsizer
        boxSizer = wx.BoxSizer( wx.VERTICAL )

        #  staticboxSizer for variable select
        sbSizer_select = wx.StaticBoxSizer( wx.StaticBox( self, wx.ID_ANY, u"Select Variable" ), wx.VERTICAL )

        bSizer1 = wx.BoxSizer( wx.HORIZONTAL )

        bSizer_unselected = wx.BoxSizer( wx.VERTICAL )

        bSizer_choose = wx.BoxSizer( wx.HORIZONTAL )

        self.staticText_unselected = wx.StaticText( sbSizer_select.GetStaticBox(), wx.ID_ANY, u"Unselected Variables", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText_unselected.Wrap( -1 )
        bSizer_choose.Add( self.staticText_unselected, 1, wx.ALIGN_CENTER|wx.ALL, 5 )

        self.textCtrl_search = wx.TextCtrl( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer_choose.Add( self.textCtrl_search, 1, wx.ALIGN_RIGHT|wx.ALL, 5 )

        bSizer_unselected.Add( bSizer_choose, 0, wx.EXPAND, 5 )

        listBox_selectChoices = []
        self.listBox_unselected = wx.ListBox( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, listBox_selectChoices, wx.LB_MULTIPLE )
        bSizer_unselected.Add( self.listBox_unselected, 1, wx.ALIGN_CENTER|wx.ALL|wx.EXPAND, 5 )

        bSizer1.Add( bSizer_unselected, 2, wx.EXPAND, 5 )

        bSizer_selectbutton = wx.BoxSizer( wx.VERTICAL )

        bSizer_button0 = wx.BoxSizer( wx.VERTICAL )

        self.staticText_empty = wx.StaticText( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText_empty.Wrap( -1 )
        bSizer_button0.Add( self.staticText_empty, 0, wx.ALIGN_CENTER|wx.ALL|wx.EXPAND, 5 )

        bSizer_selectbutton.Add( bSizer_button0, 1, wx.ALIGN_CENTER|wx.ALL, 5 )

        bSizer_select = wx.BoxSizer( wx.VERTICAL )

        self.bpButton_select = wx.BitmapButton( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.Bitmap( u"arrow_forward_16px.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, wx.BU_AUTODRAW )
        bSizer_select.Add( self.bpButton_select, 0, wx.ALIGN_CENTER|wx.ALL, 5 )

        self.bpButton_selectall = wx.BitmapButton( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.Bitmap( u"double_arrow_forward_16px.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, wx.BU_AUTODRAW )
        bSizer_select.Add( self.bpButton_selectall, 0, wx.ALIGN_CENTER|wx.ALL, 5 )

        bSizer_selectbutton.Add( bSizer_select, 1, wx.ALIGN_CENTER|wx.ALL|wx.EXPAND, 5 )

        bSizer_unselect = wx.BoxSizer( wx.VERTICAL )

        self.bpButton_unselect = wx.BitmapButton( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.Bitmap( u"arrow_back_16px.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, wx.BU_AUTODRAW )
        bSizer_unselect.Add( self.bpButton_unselect, 0, wx.ALIGN_CENTER|wx.ALL, 5 )

        self.bpButton_unselectall = wx.BitmapButton( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.Bitmap( u"double_arrow_back_16px.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, wx.BU_AUTODRAW )
        self.bpButton_unselectall.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_WINDOW ) )

        bSizer_unselect.Add( self.bpButton_unselectall, 0, wx.ALIGN_CENTER|wx.ALL, 5 )

        bSizer_selectbutton.Add( bSizer_unselect, 1, wx.ALIGN_CENTER|wx.ALL|wx.EXPAND, 5 )

        bSizer1.Add( bSizer_selectbutton, 1, wx.EXPAND, 5 )

        bSizer_selected = wx.BoxSizer( wx.VERTICAL )

        self.staticText_selected = wx.StaticText( sbSizer_select.GetStaticBox(), wx.ID_ANY, u"Selected Variables", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText_selected.Wrap( -1 )
        bSizer_selected.Add( self.staticText_selected, 0, wx.ALL, 5 )

        listBox_selectedChoices = []
        self.listBox_selected = wx.ListBox( sbSizer_select.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, listBox_selectedChoices, wx.LB_MULTIPLE )
        bSizer_selected.Add( self.listBox_selected, 1, wx.ALIGN_CENTER|wx.ALL|wx.EXPAND, 5 )

        bSizer1.Add( bSizer_selected, 2, wx.EXPAND, 5 )

        sbSizer_select.Add( bSizer1, 1, wx.EXPAND, 5 )

        boxSizer.Add( sbSizer_select, 1, wx.ALIGN_CENTER|wx.EXPAND, 5 )

        # define staticboxsizer for export
        sbSizer_export = wx.StaticBoxSizer( wx.StaticBox( self, wx.ID_ANY, u"Export Data" ), wx.VERTICAL )

        bSizer2 = wx.BoxSizer( wx.HORIZONTAL )

        self.button_clear = wx.Button( sbSizer_export.GetStaticBox(), wx.ID_ANY, u"Clear Variable", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer2.Add( self.button_clear, 1, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

        self.button_export_mdf = wx.Button( sbSizer_export.GetStaticBox(), wx.ID_ANY, u"Mdf file export", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer2.Add( self.button_export_mdf, 1, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

        self.button_export_txt = wx.Button( sbSizer_export.GetStaticBox(), wx.ID_ANY, u"TXT and Excel file export", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer2.Add( self.button_export_txt, 1, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

        sbSizer_export.Add( bSizer2, 1, wx.EXPAND, 5 )

        boxSizer.Add( sbSizer_export, 0, wx.EXPAND, 5 )

        # define staticboxsizer for rank
        sbSizer_rank = wx.StaticBoxSizer( wx.StaticBox( self, wx.ID_ANY, u"Rank Variable" ), wx.VERTICAL )

        bSizer3 = wx.BoxSizer( wx.HORIZONTAL )

        self.button_config = wx.Button( sbSizer_rank.GetStaticBox(), wx.ID_ANY, u"Read Rank_Config file", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer3.Add( self.button_config, 1, wx.ALIGN_CENTER|wx.ALL, 5 )

        self.button_txtrank = wx.Button( sbSizer_rank.GetStaticBox(), wx.ID_ANY, u"TXT file rank", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer3.Add( self.button_txtrank, 1, wx.ALIGN_CENTER|wx.ALL, 5 )

        self.button_excelrank = wx.Button( sbSizer_rank.GetStaticBox(), wx.ID_ANY, u"EXCEL file rank", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer3.Add( self.button_excelrank, 1, wx.ALIGN_CENTER|wx.ALL, 5 )

        sbSizer_rank.Add( bSizer3, 0, wx.EXPAND, 5 )

        boxSizer.Add( sbSizer_rank, 0, wx.EXPAND, 5 )

        # define staticboxsizer for resample

        sbSizer_resample = wx.StaticBoxSizer( wx.StaticBox( self, wx.ID_ANY, u"OBD Data Process" ), wx.VERTICAL )

        bSizer4 = wx.BoxSizer( wx.HORIZONTAL )

        self.button_config_resample = wx.Button( sbSizer_resample.GetStaticBox(), wx.ID_ANY, u"Read Resample_Config file", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer4.Add( self.button_config_resample, 1, wx.ALIGN_CENTER, 5 )

        self.button_mdfresample_config = wx.Button( sbSizer_resample.GetStaticBox(), wx.ID_ANY, u"MDF Data Resample", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer4.Add( self.button_mdfresample_config, 1, wx.ALIGN_CENTER|wx.ALL, 5 )

        self.button_OBD_Fleet = wx.Button( sbSizer_resample.GetStaticBox(), wx.ID_ANY, u"OBD Fleet Data Process", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer4.Add( self.button_OBD_Fleet, 1, wx.ALIGN_CENTER, 5 )

        sbSizer_resample.Add( bSizer4, 0, wx.EXPAND, 5 )

        boxSizer.Add( sbSizer_resample, 0, wx.EXPAND, 5 )

        self.SetSizer( boxSizer )
        self.Layout()

        #******************************  define menubar *********************************#

        self.menubar = wx.MenuBar( 0 )

        # define file_menu
        self.file_menu = wx.Menu()
        self.read_excelfile = wx.MenuItem( self.file_menu, wx.ID_ANY, u"Read excel", wx.EmptyString, wx.ITEM_NORMAL )
        self.file_menu.Append( self.read_excelfile )

        self.read_txtfile = wx.MenuItem( self.file_menu, wx.ID_ANY, u"Read txt", wx.EmptyString, wx.ITEM_NORMAL )
        self.file_menu.Append( self.read_txtfile )

        self.read_mdffile = wx.MenuItem( self.file_menu, wx.ID_ANY, u"Read mdf", wx.EmptyString, wx.ITEM_NORMAL )
        self.file_menu.Append( self.read_mdffile )

        self.menubar.Append( self.file_menu, u"File" )

        # define export_menu
        self.export_menu = wx.Menu()
        self.export_mdf_excelfile = wx.MenuItem( self.export_menu, wx.ID_ANY, u"Export mdf to excel", wx.EmptyString, wx.ITEM_NORMAL )
        self.export_menu.Append( self.export_mdf_excelfile )
        self.export_txt_excelfile = wx.MenuItem( self.export_menu, wx.ID_ANY, u"Export txt and excel to excel", wx.EmptyString, wx.ITEM_NORMAL )
        self.export_menu.Append( self.export_txt_excelfile )

        self.menubar.Append( self.export_menu, u"Export" )

        # define rank_menu
        self.rank_menu = wx.Menu()
        self.read_rankconfigfile = wx.MenuItem( self.rank_menu, wx.ID_ANY, u"Read rank_config file", wx.EmptyString, wx.ITEM_NORMAL )
        self.rank_menu.Append( self.read_rankconfigfile )

        self.rank_txt = wx.MenuItem( self.rank_menu, wx.ID_ANY, u"TXT file rank", wx.EmptyString, wx.ITEM_NORMAL )
        self.rank_menu.Append( self.rank_txt )

        self.rank_excel = wx.MenuItem( self.rank_menu, wx.ID_ANY, u"EXCEL file rank", wx.EmptyString, wx.ITEM_NORMAL )
        self.rank_menu.Append( self.rank_excel )

        self.menubar.Append( self.rank_menu, u"Rank" )

        # define resample_menu
        self.resample_menu = wx.Menu()
        self.read_resample_configfile = wx.MenuItem( self.rank_menu, wx.ID_ANY, u"Read resample_config file", wx.EmptyString, wx.ITEM_NORMAL )
        self.resample_menu.Append( self.read_resample_configfile )

        self.resample_mdf_config = wx.MenuItem( self.rank_menu, wx.ID_ANY, u"MDF data resample", wx.EmptyString, wx.ITEM_NORMAL )
        self.resample_menu.Append( self.resample_mdf_config )

#        self.resample_mdf_all = wx.MenuItem( self.rank_menu, wx.ID_ANY, u"MDF data resample (all variable)", wx.EmptyString, wx.ITEM_NORMAL )
#        self.rank_menu.Append( self.resample_mdf_all )

        self.menubar.Append( self.resample_menu, u"Resample" )

        # define OBD module
        self.OBD_menu = wx.Menu()
        self.OBD_Fleet = wx.MenuItem( self.rank_menu, wx.ID_ANY, u"OBD Fleet Data Process", wx.EmptyString, wx.ITEM_NORMAL )
        self.OBD_menu.Append( self.OBD_Fleet )

        self.menubar.Append( self.OBD_menu, u"OBD" )

        # define help document
        self.help_menu = wx.Menu()
        self.help_txt = wx.MenuItem( self.help_menu, wx.ID_ANY, u"Help", wx.EmptyString, wx.ITEM_NORMAL )
        self.help_menu.Append( self.help_txt )

        self.about_program = wx.MenuItem( self.help_menu, wx.ID_ANY, u"About...", wx.EmptyString, wx.ITEM_NORMAL )
        self.help_menu.Append( self.about_program )

        self.menubar.Append( self.help_menu, u"Help" )

        self.SetMenuBar( self.menubar )

        # define statusbar
        self.statusBar = self.CreateStatusBar()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl_search.Bind( wx.EVT_TEXT, self.var_search )

        self.bpButton_select.Bind( wx.EVT_BUTTON, self.select )
        self.bpButton_selectall.Bind( wx.EVT_BUTTON, self.select_all )
        self.bpButton_unselect.Bind( wx.EVT_BUTTON, self.unselect )
        self.bpButton_unselectall.Bind( wx.EVT_BUTTON, self.unselect_all )

        self.button_clear.Bind( wx.EVT_BUTTON, self.clear_selection )
        self.button_export_mdf.Bind( wx.EVT_BUTTON, self.export_mdf_to_excel )
        self.button_export_txt.Bind( wx.EVT_BUTTON, self.export_txt_to_excel )

        self.button_config.Bind( wx.EVT_BUTTON, self.read_config )
        self.button_txtrank.Bind( wx.EVT_BUTTON, self.txt_rank )
        self.button_excelrank.Bind( wx.EVT_BUTTON, self.excel_rank )

        self.button_config_resample.Bind(wx.EVT_BUTTON, self.read_config)
        self.button_mdfresample_config.Bind(wx.EVT_BUTTON, self.mdf_resample)
        self.button_OBD_Fleet.Bind(wx.EVT_BUTTON, self.OBD_Fleet_Process)

        self.Bind( wx.EVT_MENU, self.read_excel, id = self.read_excelfile.GetId() )
        self.Bind( wx.EVT_MENU, self.read_txt, id = self.read_txtfile.GetId() )
        self.Bind( wx.EVT_MENU, self.read_mdf, id = self.read_mdffile.GetId() )

        self.Bind( wx.EVT_MENU, self.export_mdf_to_excel, id = self.export_mdf_excelfile.GetId() )
        self.Bind( wx.EVT_MENU, self.export_txt_to_excel, id = self.export_txt_excelfile.GetId() )

        self.Bind( wx.EVT_MENU, self.read_config, id = self.read_rankconfigfile.GetId() )
        self.Bind( wx.EVT_MENU, self.txt_rank, id = self.rank_txt.GetId() )
        self.Bind( wx.EVT_MENU, self.excel_rank, id = self.rank_excel.GetId() )

        self.Bind( wx.EVT_MENU, self.read_config, id = self.read_resample_configfile.GetId() )
        self.Bind( wx.EVT_MENU, self.mdf_resample, id = self.resample_mdf_config.GetId() )

        self.Bind( wx.EVT_MENU, self.OBD_Fleet_Process, id = self.OBD_Fleet.GetId() )

        self.Bind( wx.EVT_MENU, self.program_help, id = self.help_txt.GetId() )
        self.Bind( wx.EVT_MENU, self.program_about, id = self.about_program.GetId() )


    def __del__( self ):
        pass

    # Virtual event handlers, overide them in your derived class
    def var_search( self, event ):
        if list(self.listBox_unselected.GetStrings()) == []:
            return
        else:
            index_select = list(self.listBox_unselected.GetSelections())
            for i in range(len(index_select)):
                self.listBox_unselected.Deselect(index_select[i]) # clear selection
            if self.textCtrl_search.GetValue() == '':
                return
            else:
                var_search = self.textCtrl_search.GetValue()
                var_unselect = self.listBox_unselected.GetStrings()
                for i in range(len(var_unselect)):
                    if re.search(var_search,var_unselect[i],flags = re.I) != None:
                        self.listBox_unselected.SetStringSelection(var_unselect[i])

    def read_txt( self, event ):
        wildcard = 'TXT Files (*.txt)|*.txt'
        dlg0 = wx.FileDialog(None,message = 'Choose a TXT file',
                                             defaultDir = os.getcwd(),
                                             defaultFile="",
                                             wildcard = wildcard,
                                             style = wx.FD_OPEN)
        if dlg0.ShowModal() == wx.ID_OK:
            self.file_path = dlg0.GetPaths() #include path and filename
        else:
            return

        with open(self.file_path[0], 'r', errors='ignore') as f:
            strb = f.read()
        with open(self.file_path[0], 'w') as a:
            a.write(strb)

        self.rawdata = pd.read_csv(self.file_path[0], sep = '\t',encoding = 'ANSI')

        key = self.rawdata.columns

        self.listBox_unselected.Clear()
        self.listBox_selected.Clear()

        for i in range(len(key)):
            self.listBox_unselected.Append(key[i])


    def read_excel( self, event ):
        wildcard = 'excel workbook(*.xlsx)|*.xlsx|excel 97-2003 workbook(*.xls)|*.xls'
        dlg0 = wx.FileDialog(None,message = 'Choose a EXCEL file',
                                             defaultDir = os.getcwd(),
                                             defaultFile="",
                                             wildcard = wildcard,
                                             style = wx.FD_OPEN)
        if dlg0.ShowModal() == wx.ID_OK:
                self.file_path = dlg0.GetPaths() #include path and filename
        else:
            return

        self.rawdata = pd.read_excel(self.file_path[0])

        key = self.rawdata.columns

        self.listBox_unselected.Clear()
        self.listBox_selected.Clear()

        for i in range(len(key)):
            self.listBox_unselected.Append(key[i])


    def read_mdf( self, event ):
        wildcard = 'INCA Files (*.dat)|*.dat'
        dlg0 = wx.FileDialog(None,message = 'Choose a MDF file',
                                         defaultDir = os.getcwd(),
                                         defaultFile="",
                                         wildcard = wildcard,
                                         style = wx.FD_OPEN)
        if dlg0.ShowModal() == wx.ID_OK:
            self.file_path = dlg0.GetPaths()
        else:
            return

        self.rawdata = mdfreader.Mdf(self.file_path[0])

        key = list(self.rawdata.keys())

        self.listBox_unselected.Clear()
        self.listBox_selected.Clear()

        for i in range(len(key)):
            self.listBox_unselected.Append(key[i])


    def select( self, event ):
        index_select = self.listBox_unselected.GetSelections()
        str_select = []
        for i in range(len(index_select)):
            str_select.append(self.listBox_unselected.GetString(index_select[i]))
        str_len = len(str_select)
        for i in range(str_len):
            self.listBox_selected.Append(str_select[i])
            self.listBox_unselected.Delete(index_select[str_len - 1 - i]) # Reverse order to delete

    def select_all( self, event ):
        str_all = self.listBox_unselected.GetStrings()
        for i in range(len(str_all)):
            self.listBox_selected.Append(str_all[i])
        self.listBox_unselected.Clear()

    def unselect( self, event ):
        index_select = self.listBox_selected.GetSelections()
        str_select = []
        for i in range(len(index_select)):
            str_select.append(self.listBox_selected.GetString(index_select[i]))
        str_len = len(str_select)
        for i in range(str_len):
            self.listBox_unselected.Append(str_select[i])
            self.listBox_selected.Delete(index_select[str_len - 1 - i]) # Reverse order to delete

    def unselect_all( self, event ):
        str_all = self.listBox_selected.GetStrings()
        for i in range(len(str_all)):
            self.listBox_unselected.Append(str_all[i])
        self.listBox_selected.Clear()

    def clear_selection( self, event ):
        self.listBox_selected.Clear()
        self.listBox_unselected.Clear()

    def export_mdf_to_excel( self, event ):
        var_list = list(self.listBox_selected.GetStrings())
        var_len = len(var_list)
        if var_len == 0:
            return

        time_data = []
        var_data = []
        for j in range(var_len):
            try:
                time_data.append(self.rawdata[self.rawdata[var_list[j]]['master']]['data'])
                var_data.append(self.rawdata[var_list[j]]['data'])
            except KeyError:
                dlg = wx.MessageDialog(None,message = 'please import mdf file',caption = 'Message')
                if dlg.ShowModal() == wx.ID_OK:
                    return

        time_start = []
        time_end = []
        for j in range(var_len):
            time_start.append(time_data[j][0])
            time_end.append(time_data[j][-1])

        # data resample

        time_max = min(time_end)
        time_min = max(time_start)
        time_sample = [int(time_min) + 1,int(time_max) - 1]

        inter = []
        data_resample = []

        interval = int((time_sample[1] - time_sample[0])/0.1 + 1) # resample according to the rate of 100ms

        time_resample = np.linspace(time_sample[0],time_sample[1],num = interval) # Resampling target sequences

        for j in range(var_len):
            inter.append(interp1d(time_data[j],var_data[j],kind='linear'))

        for j in range(var_len):
            data_resample.append(inter[j](time_resample))

        #writing excel file

        # python global variable dictionary access
        names = globals()

        names['dict_data'] = {}

        num = self.file_path[0].rfind('\\')
        file_name = self.file_path[0][num + 1:-4] # delete '.dat'


        names['dict_data']['time_resample'] = time_resample
        for j in range(var_len):
            names['dict_data'][var_list[j]] = data_resample[j]
        df = pd.DataFrame(names['dict_data'])

        path1 = 'excel_export\\'
        isExists = os.path.exists(path1)
        if not isExists:
            os.mkdir(path1)
        else:
            pass

        try:
            wb_new = Workbook()
            wb_new.save('excel_export\\%s.xlsx' % file_name)
            writer = pd.ExcelWriter('excel_export\\%s.xlsx' % file_name)
        except PermissionError:
            dlg = wx.MessageDialog(None,message = 'please close the excel output file before writing',caption = 'Message')
            if dlg.ShowModal() == wx.ID_OK:
                return

        df.to_excel(writer,sheet_name = 'selected variables',startcol = 0,index=False)
        writer.save()

        dlg = wx.MessageDialog(None,message = 'Export Excel File Complete',caption = 'Message')
        #dlg.ShowModal() : wx.ID_YES, wx.ID_NO, wx.ID_CANCEL, wx.ID_OK
        if dlg.ShowModal() == wx.ID_OK:
            return

    def export_txt_to_excel( self, event ):
        var_list = list(self.listBox_selected.GetStrings())
        var_len = len(var_list)
        if var_len == 0:
            return

        path1 = 'excel_export\\'
        isExists = os.path.exists(path1)
        if not isExists:
            os.mkdir(path1)
        else:
            pass
        num1 = self.file_path[0].rfind('\\')
        num2 = self.file_path[0].rfind('.')
        if self.file_path[0][num2 + 1:] == 'xlsx':
            excel_name = self.file_path[0][num1 + 1:-5] #delete '.xlsx'
        else:
            excel_name = self.file_path[0][num1 + 1:-4] #delete '.xls' or 'txt'

        try:
            self.rawdata.to_excel("excel_export\\%s.xlsx" % excel_name,encoding="utf_8",index= False ,columns = var_list)
        except AttributeError:
            dlg = wx.MessageDialog(None,message = 'please import txt or excel file',caption = 'Message')
            if dlg.ShowModal() == wx.ID_OK:
                return
        except PermissionError:
            dlg = wx.MessageDialog(None,message = 'please close the excel output file before writing',caption = 'Message')
            if dlg.ShowModal() == wx.ID_OK:
                return

        dlg = wx.MessageDialog(None,message = 'Export Excel File Complete',caption = 'Message')
        #dlg.ShowModal() : wx.ID_YES, wx.ID_NO, wx.ID_CANCEL, wx.ID_OK
        if dlg.ShowModal() == wx.ID_OK:
            return


    def read_config( self, event ):
        wildcard = 'Excel Workbook(*.xlsx)|*.xlsx|Excel 97-2003 Workbook(*.xls)|*.xls'
        dlg = wx.FileDialog(None,message = 'Choose Config file',
                             defaultDir = os.getcwd(),
                             defaultFile="",
                             wildcard = wildcard,
                             style = wx.FD_OPEN)
        if dlg.ShowModal() == wx.ID_OK:
            self.config_path = dlg.GetPaths() #include path and filename
        else:
            return

    def mdf_resample( self, event ):
        try:
            wb = load_workbook(self.config_path[0])
        except AttributeError:
            dlg1 = wx.MessageDialog(None,message = 'Please read config file first',caption = 'Message')
            if dlg1.ShowModal() == wx.ID_OK:
                return

        try:
            ws = wb['variable']    #the required resample variables is in sheet named 'variable'
        except KeyError:
            dlg = wx.MessageDialog(None,message = 'Please read mdf config file',caption = 'Message')
            if dlg.ShowModal() == wx.ID_OK:
                return

        sample_rate = ws['A2'].value / 1000 #required resample rate
        #    column_max = ws.max_column
        row_max = ws.max_row


        var_tuple = ws['B2':'B%d' % row_max]
        var_value = []
        for i in range(len(var_tuple)):
            var_value.append(var_tuple[i][0].value)

        #delete None data in list
        var_value = [i for i in var_value if i != None]
        var_list = list(set(var_value))  #    delete duplicated data,but change the order
        var_list.sort(key = var_value.index)  # var_list is ordered by var_value

        var_len = len(var_list)

        # get the filepath
        wildcard = 'INCA Files (*.dat)|*.dat|All Files(*.*)|*.*'
        dlg0 = wx.FileDialog(None,message = 'Choose MDF files',
                                         defaultDir = os.getcwd(),
                                         defaultFile="",
                                         wildcard = wildcard,
                                         style = wx.FD_OPEN|wx.FD_MULTIPLE)
        if dlg0.ShowModal() == wx.ID_OK:
            file_path = dlg0.GetPaths() #include filepath and filename
        #    file_fullname = dlg0.GetFilename() #only include filename
        else:
            return

        file_num = len(file_path)

        file_name = []
        for i in range(file_num):
            num = file_path[i].rfind('\\')
            file_name.append(file_path[i][num + 1:-4]) #delete '.dat'

        def newlist(num):
            newlist = []
            for i in range(num):
                newlist.append([])
            return newlist

        # read MDF file

        rawdata = []
        time_data = newlist(file_num)
        time_start = newlist(file_num)
        time_end = newlist(file_num)
        var_data = newlist(file_num)



        for i in range(file_num):
            rawdata.append(mdfreader.Mdf(file_path[i]))


        for i in range(file_num):
            for j in range(var_len):

                rawdata_varlist = list(rawdata[i].keys())
                rawdata_varlist_len = len(rawdata_varlist)
                for k in range(rawdata_varlist_len):
                    match_check = re.match(var_list[j],rawdata_varlist[k])
                    if match_check != None:
                        match_end = match_check.end()
                        if (rawdata_varlist[k][match_end:] in [':CAN1',':CCP1',':XCP1','']):
                            time_data[i].append(rawdata[i][rawdata[i][rawdata_varlist[k]]['master']]['data'])
                            var_data[i].append(rawdata[i][rawdata_varlist[k]]['data'])
                            break
                    if k == rawdata_varlist_len - 1:
                       dlg = wx.MessageDialog(None,'File %s has no variable named \"%s\"'  % (file_name[i],var_list[j]),'Error',wx.ICON_ERROR|wx.OK)
                       if dlg.ShowModal() == wx.ID_OK:
                           return

        for i in range(file_num):
            for j in range(var_len):
                time_start[i].append(time_data[i][j][0])
                time_end[i].append(time_data[i][j][-1])

        #data resample

        time_max = []
        time_min = []
        time_sample = []
        for i in range(file_num):
            time_max.append(min(time_end[i]))
            time_min.append(max(time_start[i]))
        for i in range(file_num):
            time_sample.append([int(time_min[i]) + 1,int(time_max[i]) - 1])

        inter = newlist(file_num)
        interval = []
        time_resample = []
        data_resample = newlist(file_num)

        for i in range(file_num):
            interval.append(int((time_sample[i][1] - time_sample[i][0])/sample_rate + 1)) #resample according to the required rate

        for i in range(file_num):
            time_resample.append(np.linspace(time_sample[i][0],time_sample[i][1],num = interval[i])) #Resampling target sequences

        for i in range(file_num):
            for j in range(var_len):
                inter[i].append(interp1d(time_data[i][j],var_data[i][j],kind='linear'))

        for i in range(file_num):
            for j in range(var_len):
                data_resample[i].append(inter[i][j](time_resample[i]))


        #writing excel file

        #dir_cur = os.getcwd()
        #    dict_data = [{}] * file_num
        df = []

        # python global variable dictionary access
        names = globals()

        for i in range(file_num):
            names['dict_data%d' % i] = {}

#        for i in range(file_num):
#            num = file_path[i].rfind('\\')
#            file_name.append(file_path[i][num + 1:-4]) #delete '.dat'

        for i in range(file_num):
            names['dict_data%d' % i]['time_resample'] = time_resample[i]
            for j in range(var_len):
                names['dict_data%d' % i][var_list[j]] = data_resample[i][j]
        for i in range(file_num):
            df.append(pd.DataFrame(names['dict_data%d' % i]))

        path1 = 'excel_resample\\'
        isExists = os.path.exists(path1)
        if not isExists:
            os.mkdir(path1)
        else:
            pass

        for i in range(file_num):
            try:
                wb_new = Workbook()
                wb_new.save('excel_resample\\%s.xlsx' % file_name[i])
                writer = pd.ExcelWriter('excel_resample\\%s.xlsx' % file_name[i])
                df[i].to_excel(writer,sheet_name = 'resampled data',startcol = 0,index=False)
                writer.save()
            except PermissionError:
                dlg = wx.MessageDialog(None,message = 'please close the excel output file before writing',caption = 'Message')
                if dlg.ShowModal() == wx.ID_OK:
                    return
        #writing excel file

        dlg = wx.MessageDialog(None,message = 'Resample Complete',caption = 'Message')
        #dlg.ShowModal() : wx.ID_YES, wx.ID_NO, wx.ID_CANCEL, wx.ID_OK
        if dlg.ShowModal() == wx.ID_OK:
            return

    def OBD_Fleet_Process( self, event ):
        # ********************   read Config information  ********************** #
        try:
            wb = load_workbook(self.config_path[0])
        except AttributeError:
            dlg = wx.MessageDialog(None,message = 'Please read config file first',caption = 'Message')
            if dlg.ShowModal() == wx.ID_OK:
                return

        ws = wb['trigger']   # trigger

        row_max = ws.max_row
        column_max = ws.max_column

#        var_col = np.arange(1,column_max,2)
#        trigger_col = np.arange(2,column_max + 1,2)

        tri_num = int(column_max / 2)

        cols = list(ws.columns)  # ws.columns or ws.iter_cols() is a Worksheet._cells_by_col generator

        def newlist(num):
            newlist = []
            for i in range(num):
                newlist.append([])
            return newlist

        var_list = newlist(tri_num)
        triggerlist = newlist(tri_num)

        for i in range(tri_num):
            for j in range(1,row_max):
                var_list[i].append(cols[2*i][j].value)
                triggerlist[i].append(cols[2*i + 1][j].value)

        #delete empty data
        for i in range(tri_num):
            var_list[i] = [j for j in var_list[i] if j != None]
            triggerlist[i] = [j for j in triggerlist[i] if j != None]

        varlist = newlist(tri_num)
        #delete duplicated variables
        for i in range(tri_num):
            varlist[i] = list(set(var_list[i]))
            varlist[i].sort(key = var_list[i].index)

        for i in range(tri_num):
            varlist[i].insert(0,'file_name')

        trigger = []
        tri_start = []
        tri_end = []
        tri_target = []
        for i in range(tri_num):
            trigger.append(triggerlist[i][0])
            tri_start.append(triggerlist[i][1])
            tri_end.append(triggerlist[i][2])
            tri_target.append(triggerlist[i][3])

        ######################## special trigger addition #############################

        #var_ICB1PASS = ['IC1OSCTM','ICB1BASE','ICM1EWMA','ICB1PASS','ICB1FAIL']
        #var_XFRDCT = ['XFR11FLHC','XFR11FHLC','XFR11FLHR','XFR11FHLR','XFR11FSR','XFR11SRAV','XFR11SDAV','XFRDCT']
        #var_OTD12DST = ['OTD12TTF','OTD12TTC','OTD12MXC','OTD12MXK','OTP12TTF']

        #  ***************************** Read Excel *******************************   #

        wildcard = 'EXCEL Files (*.xlsx)|*.xlsx|All Files(*.*)|*.*'
        dlg0 = wx.FileDialog(None,message = 'Choose Resampled EXCEL files',
                                         defaultDir = r'\excel_resample',
                                         defaultFile="",
                                         wildcard = wildcard,
                                         style = wx.FD_OPEN|wx.FD_MULTIPLE)

        if dlg0.ShowModal() == wx.ID_OK:
            file_path = dlg0.GetPaths() #include path and filename
        #    file_fullname = dlg0.GetFilename() #only include filename
        else:
            return

        file_num = len(file_path)

        file_name = []
        for i in range(file_num):
            num = file_path[i].rfind('\\')
            file_name.append(file_path[i][num + 1:-5]) #delete .xlsx

        df_raw = []

        for i in range(file_num):
            df_raw.append(pd.read_excel(file_path[i]))

        var_len = []
        for i in range(file_num):
            var_len.append(len(df_raw[i].iloc[:,1]))

        for i in range(file_num):
            df_raw[i]['file_name'] = [file_name[i]] * var_len[i]

        def newdict(num):
            newlist = []
            for i in range(num):
                newlist.append({})
            return newlist

        # tri_index is a list composed by dictionaries
        tri_index = newdict(file_num)

        for i in range(file_num):
            for j in range(tri_num):
                tri_index[i]['trigger%d' % (j + 1)] = []

        # df_filter is a combined list based on file number and
        df_filter = newlist(file_num)

        for i in range(file_num):
            for j in range(tri_num):
                try:
                    df_filter[i].append(df_raw[i][varlist[j]])
                except KeyError:
                    dlg = wx.MessageDialog(None,message = 'Please read resampled file',caption = 'Message')
                    if dlg.ShowModal() == wx.ID_OK:
                        return

        #  ****************************** Filter Data ******************************   #

        #judging criteria
        for i in range(file_num):
            length = len(df_raw[i].iloc[:,1])
            for j in range(tri_num):
                for k in range(1,length - 1):
        #            tri_max = max(tri_start[j],tri_end[j])
        #            tri_min = min(tri_start[j],tri_end[j])

                    if (df_raw[i][trigger[j]][k] == tri_start[j]) and \
                    (df_raw[i][trigger[j]][k + 1] != tri_start[j]) and \
                    (df_raw[i][trigger[j]][k - 1] == tri_start[j]):
                        index_start = k
                        for m in range(k + 1,length - 1):
                            if (df_raw[i][trigger[j]][m] == tri_end[j]) and \
                            (df_raw[i][trigger[j]][m + 1] == tri_end[j]) and \
                            (df_raw[i][trigger[j]][m - 1] != tri_end[j]):

                                index_end = m
                                series_filter = df_raw[i][trigger[j]][k:m + 1]
                                series_diff = np.diff(series_filter,n = 1) # calculate first-order differential
                                series_max = max(series_diff)
                                series_min = min(series_diff)

                                if (series_max >= 0 and series_min <= 0) | (series_diff.all() == False):
                                    break

                                if tri_target[j] == tri_start[j]:
                                    index_target = index_start
                                else:
                                    index_target = index_end
                                tri_index[i]['trigger%d' % (j + 1)].append(index_target)
                                break

        # df_final is a 2-level list composed by filtered data
        df_final = newlist(file_num)

        for i in range(file_num):
            for j in range(tri_num):
                df_final[i].append(df_filter[i][j].iloc[tri_index[i]['trigger%d' % (j + 1)]])

        names = globals()
        for i in range(tri_num):
            names['df_concat%d' % i] = pd.DataFrame([],columns = varlist[i])  #creat a empty dataframe

        for i in range(tri_num):
            for j in range(file_num):
                names['df_concat%d' % i] = pd.concat([names['df_concat%d' % i],df_final[j][i]],ignore_index = True)

        #  *************************** Write Excel ******************************   #

        path1 = 'excel_OBD\\'
        isExists = os.path.exists(path1)
        if not isExists:
            os.mkdir(path1)
        else:
            pass

        excel_path = 'excel_OBD\\trigger.xlsx'
        if not os.path.exists(excel_path):
            wb_new = Workbook()
            wb_new.save(excel_path)
            writer = pd.ExcelWriter(excel_path)
            for i in range(tri_num):
                names['df_concat%d' % i].to_excel(writer,sheet_name = 'trigger%d' % (i + 1),index = False)
            writer.save()
        else:
            wb_read = load_workbook(excel_path)
            writer = pd.ExcelWriter(excel_path)
            for i in range(tri_num):
                ws_read = wb_read['trigger%d' % (i + 1)]
                row_max = ws_read.max_row
                names['df_read%d' % i] = pd.read_excel(excel_path,'trigger%d' % (i + 1))
                names['df_mix%d' % i] = pd.concat([names['df_read%d' % i],names['df_concat%d' % i]],ignore_index = True)
                names['df_mix%d' % i].to_excel(writer,sheet_name = 'trigger%d' % (i + 1),index = False)
            writer.save()

        dlg = wx.MessageDialog(None,message = 'OBD Fleet Process Successfully',caption = 'Message')
        #dlg.ShowModal() : wx.ID_YES, wx.ID_NO, wx.ID_CANCEL, wx.ID_OK
        if dlg.ShowModal() == wx.ID_OK:
            return

    def txt_rank( self, event ):
        try:
            wb = load_workbook(self.config_path[0])
        except AttributeError:
            dlg1 = wx.MessageDialog(None,message = 'Please read config file first',caption = 'Message')
            if dlg1.ShowModal() == wx.ID_OK:
                return

        try:
            ws = wb['rank']
        except KeyError:
            dlg = wx.MessageDialog(None,message = 'Please read rank config file',caption = 'Message')
            if dlg.ShowModal() == wx.ID_OK:
                return

        rows = []
        for row in ws.iter_rows():
            rows.append(row)

        col_max = ws.max_column

        var_target = []
        for i in range(col_max):
            var_target.append(rows[0][i].value)

        wildcard = 'TXT File(*.txt)|*.txt'
        dlg0 = wx.FileDialog(None,message = 'Choose a TXT file',
                                             defaultDir = os.getcwd(),
                                             defaultFile="",
                                             wildcard = wildcard,
                                             style = wx.FD_OPEN|wx.FD_MULTIPLE)
        if dlg0.ShowModal() == wx.ID_OK:
            file_path = dlg0.GetPaths() #include path and filename
        else:
            return

        file_num = len(file_path)

        df = []
        for i in range(file_num):
            with open(file_path[i], 'r', errors='ignore') as f:
                strb = f.read()
            with open(file_path[i], 'w') as a:
                a.write(strb)

        for i in range(file_num):
            df.append(pd.read_csv(file_path[i], sep = '\t',encoding = 'ANSI'))

        index = []
        for i in range(file_num):
            index.append(df[i].iloc[:1].size)

        NAN_list = []
        for i in range(file_num):
            for j in range(index[i]):
                NAN_list.append(np.NAN)

        for i in range(file_num):
            for j in range(len(var_target)):
                if var_target[j] not in df[i].columns:
                    df[i][var_target[j]] = NAN_list[i]

        df_filter = []
        for i in range(file_num):
            df_filter.append(df[i][var_target])

        path2 = 'excel_rank\\'
        isExists = os.path.exists(path2)
        if not isExists:
            os.mkdir(path2)
        else:
            pass


        excel_name = []
        for i in range(file_num):
            num = file_path[i].rfind('\\')
            excel_name.append(file_path[i][num + 1:-4]) #delete '.txt'

        #sort according to the sequence specified in VarList.xlsx
        for i in range(file_num):
            try:
                df_filter[i].to_excel("excel_rank\\%s.xlsx" % excel_name[i],encoding="utf_8",index=False,columns = var_target)
            except PermissionError:
                dlg = wx.MessageDialog(None,message = 'please close the excel output file before writing',caption = 'Message')
                if dlg.ShowModal() == wx.ID_OK:
                    return

        dlg2 = wx.MessageDialog(None,message = 'Excel Rank Complete',caption = 'Message')
        if dlg2.ShowModal() == wx.ID_OK:
            return


    def excel_rank( self, event ):

        try:
            wb = load_workbook(self.config_path[0])
        except AttributeError:
            dlg1 = wx.MessageDialog(None,message = 'Please read config file first',caption = 'Message')
            if dlg1.ShowModal() == wx.ID_OK:
                return

        try:
            ws = wb['rank']
        except KeyError:
            dlg = wx.MessageDialog(None,message = 'Please read rank config file',caption = 'Message')
            if dlg.ShowModal() == wx.ID_OK:
                return

        rows = []
        for row in ws.iter_rows():
            rows.append(row)

        col_max = ws.max_column

        var_target = []
        for i in range(col_max):
            var_target.append(rows[0][i].value)

        wildcard = 'Excel Workbook(*.xlsx)|*.xlsx|Excel 97-2003 Workbook(*.xls)|*.xls'
        dlg0 = wx.FileDialog(None,message = 'Choose a EXCEL file',
                                             defaultDir = os.getcwd(),
                                             defaultFile="",
                                             wildcard = wildcard,
                                             style = wx.FD_OPEN|wx.FD_MULTIPLE)
        if dlg0.ShowModal() == wx.ID_OK:
            file_path = dlg0.GetPaths() #include path and filename
        else:
            return

        file_num = len(file_path)

        df = []
        for i in range(file_num):
            df.append(pd.read_excel(file_path[i]))

        index = []
        for i in range(file_num):
            index.append(df[i].iloc[:1].size)

        NAN_list = []
        for i in range(file_num):
            for j in range(index[i]):
                NAN_list.append(np.NAN)

        for i in range(file_num):
            for j in range(len(var_target)):
                if var_target[j] not in df[i].columns:
                    df[i][var_target[j]] = NAN_list[i]

        df_filter = []
        for i in range(file_num):
            df_filter.append(df[i][var_target])

        path2 = 'excel_rank\\'
        isExists = os.path.exists(path2)
        if not isExists:
            os.mkdir(path2)
        else:
            pass

        excel_name = []
        for i in range(file_num):
            num1 = file_path[i].rfind('\\')
            num2 = file_path[i].rfind('.')
            if file_path[i][num2 + 1:] == 'xlsx':
                excel_name.append(file_path[i][num1 + 1:-5]) #delete '.xlsx'
            else:
                excel_name.append(file_path[i][num1 + 1:-4]) #delete '.xls'

        #sort according to the sequence in VarList.xlsx
        for i in range(file_num):
            try:
                df_filter[i].to_excel("excel_rank\\%s.xlsx" % excel_name[i],encoding="utf_8",index=False,columns = var_target)
            except PermissionError:
                dlg = wx.MessageDialog(None,message = 'please close the excel output file before writing',caption = 'Message')
                if dlg.ShowModal() == wx.ID_OK:
                    return

        dlg2 = wx.MessageDialog(None,message = 'Excel Rank Complete',caption = 'Message')
        if dlg2.ShowModal() == wx.ID_OK:
            return

    def program_help(self, event):
        win32api.ShellExecute(0, 'open', 'README.txt', '', '', 1)
        return

    def program_about(self, event):
        dlg = wx.MessageDialog(None,message = 'GUI Version 2.1\nAuthor: duqs\nCompany: DFTC',caption = 'About GUI')
        if dlg.ShowModal() == wx.ID_OK:
            return

if __name__ == '__main__':
    app = wx.App()
    frame = MyWindow(None)
    frame.Show()
    app.MainLoop()
    del app


