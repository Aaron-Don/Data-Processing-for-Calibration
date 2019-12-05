# -*- coding: utf-8 -*-
"""
DFTC
Created on Fri Aug 16 09:07:21 2019
MDF data resample
@author: duqs
"""
import os,sys,time

import mdfreader

import pandas as pd
import numpy as np
#import matplotlib as mpl
#import matplotlib.pyplot as plt

from scipy.interpolate import interp1d

#import tkinter as tk
#from tkinter import filedialog
import wx       #   wxPython

#import openpyxl
from  openpyxl import  Workbook 
from openpyxl  import load_workbook

# ************************************************************************************ #

class MyWindow(wx.Frame):
    def __init__(self,parent,title = 'Program Running Status',size = (350,75)):
        super(MyWindow,self).__init__(parent,title = title,size = size)
        self.panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.HORIZONTAL)
#        Wx.Font(pointsize, fontfamily, fontstyle, fontweight)
        global mes
        font = wx.Font(18, wx.ROMAN, wx.ITALIC, wx.NORMAL)
        #StaticText(parent, id=ID_ANY, label=EmptyString, pos=DefaultPosition,size=DefaultSize, style=0, name=StaticTextNameStr)
        self.stbox = wx.StaticText(self.panel,label = mes,size = (300,75),style = wx.ALIGN_CENTER)
        self.stbox.SetFont(font)
        vbox.Add(self.stbox,proportion = 1,flag = wx.ALIGN_CENTER|wx.ALL)
        
        self.panel.SetSizer(vbox)
        self.panel.Fit()
        self.Center()
         
def resample():    
    app = wx.App()
    global mes
    mes = 'Program Start'
    
    frame = MyWindow(None)
    frame.Show()
    
    path = os.getcwd()
    os.chdir(path)
    
    wb = load_workbook('Config_resample.xlsx')
    ws = wb.active
    
    sample_rate = ws['A2'].value / 1000 #需求的采样频率
    #    column_max = ws.max_column
    row_max = ws.max_row
    
    
    var_tuple = ws['B2':'B%d' % row_max]
    var_value = []
    for i in range(len(var_tuple)):
        var_value.append(var_tuple[i][0].value)
        
    #去除列表中的None值
    var_value = [i for i in var_value if i != None]
    var_len = len(var_value)
    
    #import and get required rawdata
    #root = tk.Tk()  # 创建一个Tkinter.Tk()实例
    #root.withdraw() # 将Tkinter.Tk()实例隐藏
    #获取选取的文件路径
    wildcard = 'INCA Files (*.dat)|*.dat|All Files(*.*)|*.*'
    dlg0 = wx.FileDialog(None,message = 'Choose a MDF file',
                                     defaultDir = os.getcwd(),
                                     defaultFile="",
                                     wildcard = wildcard,
                                     style = wx.FD_OPEN|wx.FD_MULTIPLE)
    if dlg0.ShowModal() == wx.ID_OK:
        file_path = dlg0.GetPaths() #包括路径和文件名
    #    file_fullname = dlg0.GetFilename() #只包括了文件名
    else:
        time.sleep(1)
        wx.Exit() #退出wxPython
        sys.exit() 
     
    file_num = len(file_path)
    
    def newlist(num):
        newlist = []
        for i in range(num):
            newlist.append([])
        return newlist
    
    #读取MDF文件    
    time.sleep(2)
    mes = 'Reading Mdf File...'
    frame.stbox.SetLabel(mes)
    
    
    rawdata = []
    time_data = newlist(file_num)
    time_start = newlist(file_num)
    time_end = newlist(file_num)
    var_data = newlist(file_num)
    
    for i in range(file_num):
        rawdata.append(mdfreader.Mdf(file_path[i]))
    for i in range(file_num):
        for j in range(var_len):
            rawdata_varlist = list(rawdata[i].keys())
            rawdata_varlist_len = len(rawdata_varlist)
            for k in range(rawdata_varlist_len):
                match_check = re.match(var_list[j], rawdata_varlist[k])
                if match_check != None:
                    match_end = match_check.end()
                    if (rawdata_varlist[k][match_end:] in [':CAN1', ':CCP1', ':XCP1', '']):
                        time_data[i].append(rawdata[i][rawdata[i][rawdata_varlist[k]]['master']]['data'])
                        var_data[i].append(rawdata[i][rawdata_varlist[k]]['data'])
                        break
                if k == rawdata_varlist_len - 1:
                    dlg = wx.MessageDialog(None, 'File %s has no variable named \"%s\"' % (file_name[i], var_list[j]),'Error', wx.ICON_ERROR | wx.OK)
                    if dlg.ShowModal() == wx.ID_OK:
                        time.sleep(1)
                        wx.Exit()
                        sys.exit() #正常退出程序

    for i in range(file_num):
        for j in range(var_len):
            time_start[i].append(time_data[i][j][0])
            time_end[i].append(time_data[i][j][-1])
    
    
    #数据重采样
    time.sleep(2)
    mes = 'Resampling Data...'
    frame.stbox.SetLabel(mes)
    
    time_max = []
    time_min = []
    time_sample = []
    for i in range(file_num):
        time_max.append(min(time_end[i]))
        time_min.append(max(time_start[i]))
    for i in range(file_num):    
        time_sample.append([int(time_min[i]) + 1,int(time_max[i]) - 1])
    
    inter = newlist(file_num)
    interval = []
    time_resample = []
    data_resample = newlist(file_num)
    
    for i in range(file_num):
        interval.append(int((time_sample[i][1] - time_sample[i][0])/sample_rate + 1)) #按要求频率重采样
    
    for i in range(file_num):    
        time_resample.append(np.linspace(time_sample[i][0],time_sample[i][1],num = interval[i])) #重采样目标时间序列
    
    for i in range(file_num):
        for j in range(var_len):
            inter[i].append(interp1d(time_data[i][j],var_data[i][j],kind='linear'))
            
    for i in range(file_num):
        for j in range(var_len):        
            data_resample[i].append(inter[i][j](time_resample[i]))
            
    
    #writing excel file
    time.sleep(2)
    mes = 'Writing Excel...'
    frame.stbox.SetLabel(mes)    
    
    file_name = []
    #dir_cur = os.getcwd()
    #    dict_data = [{}] * file_num
    df = []
    
    #python全局变量字典访问
    names = globals()
    
    for i in range(file_num):
        names['dict_data%d' % i] = {}
    
    for i in range(file_num):
        num = file_path[i].rfind('\\')
        file_name.append(file_path[i][num + 1:-4]) #去掉.dat
    #    for i in range(file_num):    
    #        ws_new.append(wb_new.create_sheet(file_name[i]))
    
    for i in range(file_num):
        names['dict_data%d' % i]['time_resample'] = time_resample[i]
        for j in range(var_len):
            names['dict_data%d' % i][var_value[j]] = data_resample[i][j]
    for i in range(file_num):
        df.append(pd.DataFrame(names['dict_data%d' % i]))        
    
    path1 = 'excel\\'
    isExists = os.path.exists(path1)
    if not isExists:
        os.mkdir(path1)
    else:
        pass
        
    for i in range(file_num):
        wb_new = Workbook()
        wb_new.save('excel\\%s.xlsx' % file_name[i])
        writer = pd.ExcelWriter('excel\\%s.xlsx' % file_name[i])  
        df[i].to_excel(writer,sheet_name = 'resampled data',startcol = 0,index=False)
        writer.save()        
    
    #writing excel file
    time.sleep(2)
    mes = 'Program End'
    frame.stbox.SetLabel(mes) 
    
    dlg = wx.MessageDialog(None,message = 'Resample Completed',caption = 'Message')
    #dlg.ShowModal() : wx.ID_YES, wx.ID_NO, wx.ID_CANCEL, wx.ID_OK
    if dlg.ShowModal() == wx.ID_OK:
        time.sleep(1)
        frame.Close()
        
    dlg.Destroy()    
    app.MainLoop()
    del app


if __name__ == '__main__':
    resample()








