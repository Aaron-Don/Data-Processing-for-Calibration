# -*- coding: utf-8 -*-
"""
Created on Wed Nov 27 16:48:42 2019
OBD Calibration data extract
@author: duqs
Company: DFTC
"""

import re

import time,os,sys

import wx

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,column_index_from_string

class MyWindow(wx.Frame):
    def __init__(self,parent,title = 'Program Running Status',size = (350,75)):
        super(MyWindow,self).__init__(parent,title = title,size = size)
        self.panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.HORIZONTAL)
#        Wx.Font(pointsize, fontfamily, fontstyle, fontweight)
        global mes
        font = wx.Font(18, wx.ROMAN, wx.ITALIC, wx.NORMAL)
        #StaticText(parent, id=ID_ANY, label=EmptyString, pos=DefaultPosition,size=DefaultSize, style=0, name=StaticTextNameStr)
        self.stbox = wx.StaticText(self.panel,label = 'Program Running',size = (300,75),style = wx.ALIGN_CENTER)
        self.stbox.SetFont(font)
        vbox.Add(self.stbox,proportion = 1,flag = wx.ALIGN_CENTER|wx.ALL)

        self.panel.SetSizer(vbox)
        self.panel.Fit()
        self.Center()

app = wx.App()
frame = MyWindow(None)
frame.Show()

path = os.getcwd()
os.chdir(path)

wb0 = load_workbook('Summary Table.xlsx')
ws_config = wb0['Config']
col_max = ws_config.max_column
search_num = int(col_max / 2)

cols_config = list(ws_config.columns)

config_list = []
for i in range(search_num):
    config_list.append({})

for i in range(search_num):
    config_list[i][cols_config[2*i][1].value] = cols_config[2*i + 1][1].value
    config_list[i][cols_config[2*i][2].value] = cols_config[2*i + 1][2].value

search_list = []
fill_loc = []

ws_para = wb0['Parameter List']
cols_para = list(ws_para.columns)

ws_table = wb0['Table']
max_col_table = ws_table.max_column
ws_table.delete_cols(1,amount = max_col_table)

for i in range(search_num):
    col_start = list(config_list[i].keys())[0]
    col_end = list(config_list[i].keys())[1]
    fill_col = config_list[i][col_start]
    fill_letter = re.search(r'[A-Za-z]+',fill_col).group()
    col_letter = re.search(r'[A-Za-z]+',col_start).group()
    col_index = column_index_from_string(col_letter)
    row_start = int(re.search(r'[0-9]+',col_start).group())
    row_end = int(re.search(r'[0-9]+',col_end).group())
    for j in range(row_start,row_end + 1):
        if cols_para[col_index - 1][j - 1].value not in (None,'——','Sample Time'):
            search_list.append(cols_para[col_index - 1][j - 1].value)
            fill_loc.append(fill_letter + str(j))
        else:
            continue

wildcard = 'excel workbook(*.xlsx)|*.xlsx|excel 97-2003 workbook(*.xls)|*.xls'
dlg0 = wx.FileDialog(None,message = 'Choose Calibration data file',
                                     defaultDir = os.getcwd(),
                                     defaultFile="",
                                     wildcard = wildcard,
                                     style = wx.FD_OPEN)
if dlg0.ShowModal() == wx.ID_OK:
    cal_path = dlg0.GetPaths()
else:
    time.sleep(1)
    wx.Exit() # exit wxPython
    sys.exit()

search_len = len(search_list)

wb1 = load_workbook(cal_path[0])
ws_cal = wb1.active

rows_search = list(ws_cal.rows)
row_max_search = ws_cal.max_row
col_max_search = ws_cal.max_column

table_num = 1
table_fill_row = 1
for i in range(search_len):
    for j in range(1,row_max_search + 1):
        if ws_cal['B' + str(j)].value == search_list[i]:
            if search_list[i][5] == 't':
                ws_para[fill_loc[i]] = 'Table' + str(table_num)
                ws_table['A' + str(table_fill_row)] = 'Table' + str(table_num)
                ws_table['C' + str(table_fill_row)] = search_list[i]
                ws_para[fill_loc[i]].hyperlink = '#Table!A' + str(table_fill_row)
                table_num += 1

                for k in range(j + 1,row_max_search):
                    none_count = 0
                    for m in range(1,col_max_search + 1):
                        col_num = get_column_letter(m)
                        if ws_cal[col_num + str(k)].value == None:
                            none_count += 1
                        else:
                            continue
                    if none_count == col_max_search:
                        end_row = k
                        break
                    else:
                        continue
                row_num = end_row - j
                for p in range(table_fill_row + 1,table_fill_row + 1 + row_num):
                    for q in range(1,col_max_search + 1):
                        col_num_table =  get_column_letter(q)
                        ws_table[col_num_table + str(p)] = ws_cal[col_num_table + str(p - table_fill_row + j)].value

                table_fill_row = table_fill_row + row_num + 1

            else:
                ws_para[fill_loc[i]] = ws_cal['C' + str(j + 2)].value

        else:
            continue

wb0.save('Summary Table.xlsx')

dlg = wx.MessageDialog(None,message = 'Calibration Data Filling Completed',caption = 'Message')

if dlg.ShowModal() == wx.ID_OK:
    time.sleep(1)
    frame.Close()

dlg.Destroy()
app.MainLoop()
del app


