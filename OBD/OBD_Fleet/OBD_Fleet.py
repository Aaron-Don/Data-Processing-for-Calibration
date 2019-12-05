# -*- coding: utf-8 -*-
"""
Created on Wed Sep 4 13:49:31 2019
OBD_Fleet data processing
@author: duqs
company: DFTC
"""

import os,sys,time
#import Mdf_resample as res

import wx

import pandas as pd 
import numpy as np

from openpyxl import Workbook 
from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter, column_index_from_string

class MyWindow(wx.Frame):
    def __init__(self,parent,title = 'Program Running Status',size = (350,75)):
        super(MyWindow,self).__init__(parent,title = title,size = size)
        self.panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.HORIZONTAL)
#        Wx.Font(pointsize, fontfamily, fontstyle, fontweight)
        global mes
        font = wx.Font(18, wx.ROMAN, wx.ITALIC, wx.NORMAL)
        #StaticText(parent, id=ID_ANY, label=EmptyString, pos=DefaultPosition,size=DefaultSize, style=0, name=StaticTextNameStr)
        self.stbox = wx.StaticText(self.panel,label = mes,size = (300,75),style = wx.ALIGN_CENTER)
        self.stbox.SetFont(font)
        vbox.Add(self.stbox,proportion = 1,flag = wx.ALIGN_CENTER|wx.ALL)
        
        self.panel.SetSizer(vbox)
        self.panel.Fit()
        self.Center()

# ********************   read Config information  ********************** #
app = wx.App()
#global mes
mes = 'Program Start'
frame = MyWindow(None)
frame.Show()

path = os.getcwd()
os.chdir(path)
    
wb = load_workbook('Config_OBDFleet.xlsx')
ws = wb['trigger']   # trigger

row_max = ws.max_row
column_max = ws.max_column

var_col = np.arange(1,column_max,2)
trigger_col = np.arange(2,column_max + 1,2)

tri_num = int(column_max / 2)

#cols = list(ws.iter_cols())    
cols = list(ws.columns)  # ws.columns or ws.iter_cols() is a Worksheet._cells_by_col generator 
    
def newlist(num):
    newlist = []
    for i in range(num):
        newlist.append([])
    return newlist

var_list = newlist(tri_num)
triggerlist = newlist(tri_num)  

for i in range(tri_num):
    for j in range(1,row_max):
        var_list[i].append(cols[2*i][j].value)
        triggerlist[i].append(cols[2*i + 1][j].value)

#delete empty data        
for i in range(tri_num):
    var_list[i] = [j for j in var_list[i] if j != None]
    triggerlist[i] = [j for j in triggerlist[i] if j != None]

varlist = newlist(tri_num)
#delete duplicated variables
for i in range(tri_num):
    varlist[i] = list(set(var_list[i]))
    varlist[i].sort(key = var_list[i].index)
    
for i in range(tri_num):
    varlist[i].insert(0,'file_name')

trigger = []
tri_start = []
tri_end = []
tri_target = []
for i in range(tri_num):
    trigger.append(triggerlist[i][0])
    tri_start.append(triggerlist[i][1])
    tri_end.append(triggerlist[i][2])
    tri_target.append(triggerlist[i][3])

######################## special trigger addition #############################

#var_ICB1PASS = ['IC1OSCTM','ICB1BASE','ICM1EWMA','ICB1PASS','ICB1FAIL']
#var_XFRDCT = ['XFR11FLHC','XFR11FHLC','XFR11FLHR','XFR11FHLR','XFR11FSR','XFR11SRAV','XFR11SDAV','XFRDCT']
#var_OTD12DST = ['OTD12TTF','OTD12TTC','OTD12MXC','OTD12MXK','OTP12TTF']
    
#  ***************************** Read Excel *******************************   #

wildcard = 'EXCEL Files (*.xlsx)|*.xlsx|All Files(*.*)|*.*'
dlg0 = wx.FileDialog(None,message = 'Choose a EXCEL file',
                                 defaultDir = os.getcwd(),
                                 defaultFile="",
                                 wildcard = wildcard,
                                 style = wx.FD_OPEN|wx.FD_MULTIPLE)

time.sleep(2)
mes = 'Reading Excel File...'
frame.stbox.SetLabel(mes)  

if dlg0.ShowModal() == wx.ID_OK:
    file_path = dlg0.GetPaths() #include path and filename
#    file_fullname = dlg0.GetFilename() #only include filename
else:
    time.sleep(1)
    wx.Exit() #退出wxPython
    sys.exit() 

file_num = len(file_path)

file_name = []
for i in range(file_num):
    num = file_path[i].rfind('\\')
    file_name.append(file_path[i][num + 1:-5]) #delete .xlsx

df_raw = []

for i in range(file_num):
    df_raw.append(pd.read_excel(file_path[i])) 

var_len = []
for i in range(file_num):
    var_len.append(len(df_raw[i].iloc[:,1]))

for i in range(file_num):
    df_raw[i]['file_name'] = [file_name[i]] * var_len[i]

def newdict(num):
    newlist = []
    for i in range(num):
        newlist.append({})
    return newlist

# tri_index is a list composed by dictionaries
tri_index = newdict(file_num)
for i in range(file_num):
    for j in range(tri_num):
        tri_index[i]['trigger%d' % (j + 1)] = []

# df_filter is a combined list based on file number and     
df_filter = newlist(file_num)
for i in range(file_num):
    for j in range(tri_num):
        df_filter[i].append(df_raw[i][varlist[j]])

#  ****************************** Filter Data ******************************   #

time.sleep(2)
mes = 'Filtering Data...'
frame.stbox.SetLabel(mes)

#judging criteria
for i in range(file_num):
    length = len(df_raw[i].iloc[:,1])
    for j in range(tri_num):
        for k in range(1,length - 1):
#            tri_max = max(tri_start[j],tri_end[j])
#            tri_min = min(tri_start[j],tri_end[j])
            
            if (df_raw[i][trigger[j]][k] == tri_start[j]) and \
            (df_raw[i][trigger[j]][k + 1] != tri_start[j]) and \
            (df_raw[i][trigger[j]][k - 1] == tri_start[j]):
                index_start = k
                for m in range(k + 1,length - 1):
                    if (df_raw[i][trigger[j]][m] == tri_end[j]) and \
                    (df_raw[i][trigger[j]][m + 1] == tri_end[j]) and \
                    (df_raw[i][trigger[j]][m - 1] != tri_end[j]):
                        
                        index_end = m                        
                        series_filter = df_raw[i][trigger[j]][k:m + 1]
                        series_diff = np.diff(series_filter,n = 1) # calculate first-order differential
                        series_max = max(series_diff)
                        series_min = min(series_diff)
                        
                        if (series_max >= 0 and series_min <= 0) | (series_diff.all() == False):
                            break                            
                            
                        if tri_target[j] == tri_start[j]:
                            index_target = index_start
                        else:
                            index_target = index_end                   
                        tri_index[i]['trigger%d' % (j + 1)].append(index_target)
                        break

# df_final is a 2-level list composed by filtered data                           
df_final = newlist(file_num)

for i in range(file_num):
    for j in range(tri_num):
        df_final[i].append(df_filter[i][j].iloc[tri_index[i]['trigger%d' % (j + 1)]])

names = globals()
for i in range(tri_num):
    names['df_concat%d' % i] = pd.DataFrame([],columns = varlist[i])  #creat a empty dataframe

for i in range(tri_num):    
    for j in range(file_num):
        names['df_concat%d' % i] = pd.concat([names['df_concat%d' % i],df_final[j][i]],ignore_index = True)

#  *************************** Write Excel ******************************   #

time.sleep(2)
mes = 'Writing Output File...'
frame.stbox.SetLabel(mes) 
        
path1 = 'excel_output\\'
isExists = os.path.exists(path1)
if not isExists:
    os.mkdir(path1)
else:
    pass

excel_path = 'excel_output\\trigger.xlsx'
if not os.path.exists(excel_path):
    wb_new = Workbook()
    wb_new.save(excel_path)
    writer = pd.ExcelWriter(excel_path)  
    for i in range(tri_num):
        names['df_concat%d' % i].to_excel(writer,sheet_name = 'trigger%d' % (i + 1),index = False)
    writer.save()
else: 
    wb_read = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path)     
    for i in range(tri_num):
        ws_read = wb_read['trigger%d' % (i + 1)]
        row_max = ws_read.max_row
        names['df_read%d' % i] = pd.read_excel(excel_path,'trigger%d' % (i + 1))
        names['df_mix%d' % i] = pd.concat([names['df_read%d' % i],names['df_concat%d' % i]],ignore_index = True)
        names['df_mix%d' % i].to_excel(writer,sheet_name = 'trigger%d' % (i + 1),index = False)
    writer.save()    
    
time.sleep(2)
mes = 'Program End'
frame.stbox.SetLabel(mes) 

dlg = wx.MessageDialog(None,message = 'Program Complete',caption = 'Message')
#dlg.ShowModal() : wx.ID_YES, wx.ID_NO, wx.ID_CANCEL, wx.ID_OK
if dlg.ShowModal() == wx.ID_OK:
    time.sleep(1)
    frame.Close()
    
dlg.Destroy()

app.MainLoop()
del app
















